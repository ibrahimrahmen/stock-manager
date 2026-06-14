"""Import historic delivery/return stats per phone from Navex xlsx exports.

Two file kinds (auto-detected by the date/Motif columns):
  - "Liste des colis livrés payées"  -> each row = a DELIVERED order
  - "Liste des colis" (with 'Motif')  -> each row = a RETURN (filtered by reason)

Phone is extracted as the first 8-digit run inside the 'Nom' column.

Return reasons that count as a real customer-fault RETURN:
    Pas de réponse, Injoignable, 3 tentatives accomplies,
    Commande non conforme, A vérifier avec expéditeur
Excluded (NOT counted as retour, and NOT as anything):
    Annulé par expéditeur (= annulé), Echange, Commande double

Usage:
    python manage.py import_customer_history /path/file.xlsx [--reset]
    python manage.py import_customer_history f1.xlsx f2.xlsx
"""
import re
from django.core.management.base import BaseCommand
from inventory.models import CustomerHistory

PHONE_RE = re.compile(r"(\d{8})")

# Return motifs that count as a real customer-fault RETURN (bad outcome).
RETURN_MOTIFS = {
    "pas de réponse",
    "injoignable",
    "3 tentatives accomplies",
    "commande non conforme",
    "téléphone fermé",
    "téléphone incorrect",
    "client disponible demain",
    "client non disponible (daté)",
    "adresse incomplète",
}
# Motifs explicitly NOT counted (sender cancel / exchange / delivery-side /
# admin / unresolved-pending). These do not count against the customer.
EXCLUDED_MOTIFS = {
    "annulé par expéditeur",
    "echange",
    "commande double",
    "livreur en liste de rejet",
    "trajet inaccessible",
    "parcours non terminé",
    "colis endommagé",
    "montant incorrect",
    "a vérifier avec expéditeur",   # unresolved/pending, not a confirmed refusal
}


def _phone(nom):
    if not nom:
        return None
    m = PHONE_RE.search(str(nom))
    return m.group(1) if m else None


class Command(BaseCommand):
    help = "Import historic customer delivery/return stats from Navex xlsx exports."

    def add_arguments(self, parser):
        parser.add_argument("files", nargs="+", help="Path(s) to .xlsx export(s)")
        parser.add_argument("--reset", action="store_true",
                            help="Delete all CustomerHistory rows before importing")
        parser.add_argument("--dry-run", action="store_true",
                            help="Parse and report, but do not write to the database")

    def handle(self, *args, **opts):
        try:
            import openpyxl
        except ImportError:
            self.stderr.write("openpyxl is required. Add it to requirements.txt.")
            return

        if opts["reset"] and not opts["dry_run"]:
            n = CustomerHistory.objects.count()
            CustomerHistory.objects.all().delete()
            self.stdout.write(f"Reset: deleted {n} existing CustomerHistory rows.")

        # Accumulate in memory, then bulk-apply.
        delivered = {}   # phone -> count
        returned = {}    # phone -> count
        skipped_excluded = 0
        skipped_nophone = 0
        unknown_motifs = {}

        for path in opts["files"]:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            # Find the header row (the one whose first cell == 'Code').
            header_row = None
            headers = []
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
                if row and str(row[0]).strip().lower() == "code":
                    header_row = i
                    headers = [str(c).strip().lower() if c else "" for c in row]
                    break
            if header_row is None:
                self.stderr.write(f"{path}: no 'Code' header found, skipping.")
                continue

            is_return_file = any("motif" in h for h in headers)
            motif_idx = headers.index(next(h for h in headers if "motif" in h)) if is_return_file else None
            nom_idx = headers.index(next((h for h in headers if h == "nom"), "nom")) if "nom" in headers else 1

            kind = "RETURNS" if is_return_file else "DELIVERED"
            count_rows = 0
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not row or not row[0]:
                    continue
                count_rows += 1
                nom = row[nom_idx] if nom_idx < len(row) else None
                ph = _phone(nom)
                if not ph:
                    skipped_nophone += 1
                    continue
                if is_return_file:
                    motif = str(row[motif_idx]).strip().lower() if motif_idx is not None and row[motif_idx] else ""
                    if motif in EXCLUDED_MOTIFS:
                        skipped_excluded += 1
                        continue
                    if motif in RETURN_MOTIFS:
                        returned[ph] = returned.get(ph, 0) + 1
                    else:
                        # Unknown motif: count as a return by default but track it.
                        unknown_motifs[motif] = unknown_motifs.get(motif, 0) + 1
                        returned[ph] = returned.get(ph, 0) + 1
                else:
                    delivered[ph] = delivered.get(ph, 0) + 1

            self.stdout.write(f"{path}: {kind}, {count_rows} rows read.")

        # Report
        self.stdout.write(f"\nDistinct phones delivered: {len(delivered)}")
        self.stdout.write(f"Distinct phones returned:  {len(returned)}")
        self.stdout.write(f"Skipped (no phone): {skipped_nophone} | Skipped (excluded motif): {skipped_excluded}")
        if unknown_motifs:
            self.stdout.write("Unknown motifs (counted as retour):")
            for m, n in sorted(unknown_motifs.items(), key=lambda x: -x[1])[:10]:
                self.stdout.write(f"   {n:5d}  {m!r}")

        if opts["dry_run"]:
            self.stdout.write("\nDry-run: nothing written.")
            return

        # Apply to DB: union of all phones.
        all_phones = set(delivered) | set(returned)
        created = 0
        updated = 0
        for ph in all_phones:
            d = delivered.get(ph, 0)
            r = returned.get(ph, 0)
            obj, was_created = CustomerHistory.objects.get_or_create(phone=ph)
            # Add (so re-running with new files accumulates rather than overwrites)
            obj.historic_delivered += d
            obj.historic_returned += r
            obj.historic_total = obj.historic_delivered + obj.historic_returned
            obj.save()
            if was_created:
                created += 1
            else:
                updated += 1

        self.stdout.write(self.style.SUCCESS(
            f"\nDone. CustomerHistory: {created} created, {updated} updated, "
            f"{len(all_phones)} phones total."
        ))
